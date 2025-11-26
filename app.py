# carwash_server.py
from datetime import datetime, timedelta
import json
from flask import Flask, request, jsonify, render_template, send_from_directory, abort, send_file
import io
from openpyxl import Workbook
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from flask import redirect, url_for
import os

BASEDIR = os.path.abspath(os.path.dirname(__file__))
# Use instance folder for database (Flask convention)
INSTANCE_PATH = os.path.join(BASEDIR, 'instance')
os.makedirs(INSTANCE_PATH, exist_ok=True)
DB_PATH = os.path.join(INSTANCE_PATH, 'carwash.db')

app = Flask(__name__, static_folder='static', template_folder='static')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + DB_PATH
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)
CORS(app, resources={r"/api/*": {"origins": "*"}})

# Models
class Customer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    phone = db.Column(db.String(30))
    email = db.Column(db.String(120))
    vehicles = db.relationship('Vehicle', backref='owner', cascade='all,delete-orphan')

class Vehicle(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    reg_no = db.Column(db.String(80), nullable=False, unique=True)
    model = db.Column(db.String(120))
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'))

class Service(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    price = db.Column(db.Float, nullable=False)

class Appointment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    vehicle_id = db.Column(db.Integer, db.ForeignKey('vehicle.id'), nullable=False)
    service_id = db.Column(db.Integer, db.ForeignKey('service.id'), nullable=False)
    scheduled_at = db.Column(db.DateTime, nullable=False)
    status = db.Column(db.String(30), default='scheduled')  # scheduled/done/cancelled
    paid = db.Column(db.Boolean, default=False)
    service = db.relationship('Service')

class Payment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    appointment_id = db.Column(db.Integer, db.ForeignKey('appointment.id'), nullable=True)
    amount = db.Column(db.Float, nullable=False)
    method = db.Column(db.String(30), default='cash')
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)

# POS Sale / SaleItem
class Sale(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=True)
    vehicle_id = db.Column(db.Integer, db.ForeignKey('vehicle.id'), nullable=True)
    total = db.Column(db.Float, nullable=False)
    paid = db.Column(db.Boolean, default=True)
    method = db.Column(db.String(30), default='cash')
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    items = db.relationship('SaleItem', backref='sale', cascade='all,delete-orphan')

class SaleItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sale_id = db.Column(db.Integer, db.ForeignKey('sale.id'), nullable=False)
    service_id = db.Column(db.Integer, db.ForeignKey('service.id'), nullable=False)
    qty = db.Column(db.Integer, nullable=False)
    price = db.Column(db.Float, nullable=False)
    line_total = db.Column(db.Float, nullable=False)
    service = db.relationship('Service')

# Init DB + seed services if missing
def init_db():
    db.create_all()
    if Service.query.count() == 0:
        db.session.add_all([
            Service(name='Exterior Wash', price=150.0),
            Service(name='Full Wash + Interior', price=400.0),
            Service(name='Polish & Wax', price=700.0),
        ])
        db.session.commit()

# --- Test route to verify server is running ---
@app.route('/test')
def test():
    return jsonify({'status': 'ok', 'message': 'Server is running!'})

# --- Static pages (serve frontend) ---
@app.route('/')
def index():
    static_path = os.path.join(BASEDIR, 'static', 'index.html')
    if os.path.exists(static_path):
        return send_file(static_path)
    return send_from_directory(app.static_folder, 'index.html')

# Serve other static resources (JS/CSS) automatically by Flask's static route

# --- Serve POS frontend (static/index.html) ---
@app.route('/pos')
def pos_frontend():
    # serve the static SPA page stored in the `static` folder
    static_path = os.path.join(BASEDIR, 'static', 'index.html')
    if os.path.exists(static_path):
        return send_file(static_path)
    return send_from_directory('static', 'index.html')

# --- API endpoints (JSON) ---

# Services
@app.route('/api/services', methods=['GET', 'POST'])
def api_services():
    if request.method == 'GET':
        services = Service.query.all()
        return jsonify([{'id':s.id, 'name':s.name, 'price': s.price} for s in services])
    data = request.get_json(force=True)
    if not data or 'name' not in data or 'price' not in data:
        return jsonify({'error':'name and price required'}), 400
    s = Service(name=data['name'], price=float(data['price']))
    db.session.add(s); db.session.commit()
    return jsonify({'id': s.id}), 201

@app.route('/api/services/<int:sid>', methods=['PUT','DELETE'])
def api_service_modify(sid):
    s = Service.query.get_or_404(sid)
    if request.method == 'PUT':
        data = request.get_json(force=True)
        s.name = data.get('name', s.name)
        s.price = float(data.get('price', s.price))
        db.session.commit()
        return jsonify({'id': s.id})
    db.session.delete(s); db.session.commit()
    return jsonify({'deleted': True})

# Customers
@app.route('/api/customers', methods=['GET','POST'])
def api_customers():
    if request.method == 'GET':
        cs = Customer.query.all()
        return jsonify([{'id':c.id,'name':c.name,'phone':c.phone,'email':c.email} for c in cs])
    data = request.get_json(force=True)
    if not data or 'name' not in data:
        return jsonify({'error':'name required'}), 400
    c = Customer(name=data['name'], phone=data.get('phone'), email=data.get('email'))
    db.session.add(c); db.session.commit()
    return jsonify({'id': c.id}), 201

# Vehicles
@app.route('/api/vehicles', methods=['GET','POST'])
def api_vehicles():
    if request.method == 'GET':
        vs = Vehicle.query.all()
        return jsonify([{'id':v.id,'reg_no':v.reg_no,'model':v.model,'customer_id':v.customer_id} for v in vs])
    data = request.get_json(force=True)
    if not data or 'reg_no' not in data:
        return jsonify({'error':'reg_no required'}), 400
    v = Vehicle(reg_no=data['reg_no'], model=data.get('model'), customer_id=data.get('customer_id'))
    db.session.add(v); db.session.commit()
    return jsonify({'id': v.id}), 201

# Sales (POS) - create sale and optionally create appointment
@app.route('/api/sale', methods=['POST'])
def api_sale():
    """Accept a sale JSON from the frontend POS and persist it to the DB.
    Expected JSON shape (example):
    {
      "id": "s_xxx",
      "customer": {"id": null, "name": "Name", "phone": "..."},
      "vehicle": {"reg_no": "MH12...", "model": "..."},
      "items": [{"service_id": 1, "qty": 1, "price": 150}],
      "subtotal": 150, "tax": 0, "total": 150, "timestamp": "...", "create_appointment": true
    }
    """
    data = request.get_json() or {}
    # customer
    cust = None
    cust_in = data.get('customer') or {}
    if cust_in.get('id'):
        cust = Customer.query.get(cust_in.get('id'))
    if not cust and cust_in.get('name'):
        # try to find by name+phone
        cust = Customer.query.filter_by(name=cust_in.get('name'), phone=cust_in.get('phone')).first()
        if not cust:
            cust = Customer(name=cust_in.get('name'), phone=cust_in.get('phone'))
            db.session.add(cust); db.session.commit()

    # vehicle
    veh = None
    veh_in = data.get('vehicle') or {}
    if veh_in.get('reg_no'):
        veh = Vehicle.query.filter_by(reg_no=veh_in.get('reg_no')).first()
        if not veh:
            veh = Vehicle(reg_no=veh_in.get('reg_no'), model=veh_in.get('model'), owner=cust)
            db.session.add(veh); db.session.commit()

    # optionally create appointment
    appt = None
    if data.get('create_appointment'):
        svc_id = None
        items = data.get('items') or []
        if items:
            svc_id = int(items[0].get('service_id'))
        if svc_id and veh:
            appt = Appointment(vehicle_id=veh.id, service_id=svc_id, scheduled_at=datetime.fromisoformat(data.get('timestamp')) if data.get('timestamp') else datetime.utcnow(), status='done' if data.get('create_appointment_done') else 'scheduled', paid=bool(data.get('total')))
            db.session.add(appt); db.session.commit()

    # record payment
    if data.get('total'):
        p = Payment(appointment_id=appt.id if appt else None, amount=float(data.get('total')), method=data.get('method') or 'cash')
        db.session.add(p)
        if appt:
            appt.paid = True
        db.session.commit()

    return jsonify({'ok': True, 'customer_id': cust.id if cust else None, 'vehicle_id': veh.id if veh else None, 'appointment_id': appt.id if appt else None}), 201
@app.route('/api/sales', methods=['POST'])
def api_create_sale():
    """
    Expected JSON:
    {
      "customer": {"id":1} OR {"name":"Name","phone":"..."},
      "vehicle": {"reg_no":"MH12..","model":"..."} (optional),
      "items": [{"service_id":1,"qty":2}, ...],
      "method": "cash",
      "create_appointment": true/false,
      "appointment_service_id": <service id> (optional),
      "scheduled_at": "YYYY-MM-DD HH:MM" (optional)
    }
    """
    data = request.get_json(force=True)
    if not data or 'items' not in data or not isinstance(data['items'], list) or len(data['items'])==0:
        return jsonify({'error':'items required'}), 400

    # customer
    cust = None
    if data.get('customer'):
        cdata = data['customer']
        if isinstance(cdata, dict) and cdata.get('id'):
            cust = Customer.query.get(cdata.get('id'))
        elif isinstance(cdata, dict) and cdata.get('name'):
            cust = Customer(name=cdata.get('name'), phone=cdata.get('phone'))
            db.session.add(cust); db.session.commit()

    # vehicle
    v = None
    if data.get('vehicle') and data['vehicle'].get('reg_no'):
        v = Vehicle.query.filter_by(reg_no=data['vehicle']['reg_no']).first()
        if not v:
            v = Vehicle(reg_no=data['vehicle']['reg_no'], model=data['vehicle'].get('model'), customer_id=cust.id if cust else None)
            db.session.add(v); db.session.commit()

    # create sale
    sale = Sale(customer_id=cust.id if cust else None, vehicle_id=v.id if v else None, total=0.0, paid=True, method=data.get('method','cash'))
    db.session.add(sale); db.session.commit()

    total = 0.0
    for it in data['items']:
        svc = Service.query.get(int(it['service_id']))
        if not svc:
            db.session.rollback()
            return jsonify({'error':f'service id {it.get("service_id")} not found'}), 400
        qty = int(it.get('qty',1))
        line = svc.price * qty
        si = SaleItem(sale_id=sale.id, service_id=svc.id, qty=qty, price=svc.price, line_total=line)
        db.session.add(si)
        total += line
    sale.total = total
    db.session.commit()

    result = {'sale_id': sale.id}

    if data.get('create_appointment'):
        appt_svc = data.get('appointment_service_id') or data['items'][0]['service_id']
        scheduled_at = data.get('scheduled_at')
        if scheduled_at:
            scheduled_dt = datetime.strptime(scheduled_at, '%Y-%m-%d %H:%M')
        else:
            scheduled_dt = datetime.utcnow()
        # ensure vehicle exists (appointment needs vehicle)
        if not v:
            v = Vehicle(reg_no=f'WALKIN-{sale.id}', model='', customer_id=cust.id if cust else None)
            db.session.add(v); db.session.commit()
        appt = Appointment(vehicle_id=v.id, service_id=int(appt_svc), scheduled_at=scheduled_dt, status='done', paid=True)
        db.session.add(appt); db.session.commit()
        result['appointment_id'] = appt.id

    return jsonify(result), 201

# Get sale invoice
@app.route('/api/sales/<int:sale_id>', methods=['GET'])
def api_get_sale(sale_id):
    sale = Sale.query.get_or_404(sale_id)
    items = [{'service_id':it.service_id, 'service_name':it.service.name, 'qty':it.qty, 'price':it.price, 'line_total':it.line_total} for it in sale.items]
    return jsonify({'id':sale.id, 'customer_id':sale.customer_id, 'vehicle_id':sale.vehicle_id, 'total':sale.total, 'method':sale.method, 'timestamp':sale.timestamp.isoformat(), 'items': items})

# List sales
@app.route('/api/sales', methods=['GET'])
def api_list_sales():
    sales = Sale.query.order_by(Sale.timestamp.desc()).limit(500).all()
    out = []
    for s in sales:
        cust = Customer.query.get(s.customer_id)
        veh = Vehicle.query.get(s.vehicle_id)
        out.append({'id': s.id, 'customer': cust.name if cust else None, 'reg_no': veh.reg_no if veh else None, 'total': s.total, 'timestamp': s.timestamp.isoformat()})
    return jsonify(out)

# Simple reports
@app.route('/api/reports/daily', methods=['GET'])
def api_daily_report():
    date_str = request.args.get('date')  # YYYY-MM-DD
    if not date_str:
        return jsonify({'error':'date param required'}), 400
    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    total = db.session.query(db.func.sum(Sale.total)).filter(db.func.date(Sale.timestamp)==date_obj).scalar() or 0
    cnt = Sale.query.filter(db.func.date(Sale.timestamp)==date_obj).count()
    return jsonify({'date': date_str, 'count': cnt, 'total': total})


# --- Visual Dashboard (HTML with Chart.js) ---
@app.route('/dashboard')
def dashboard():
    cust_count = Customer.query.count()
    vehicle_count = Vehicle.query.count()
    service_count = Service.query.count()
    sale_count = Sale.query.count()

    N = 30
    today = datetime.utcnow().date()
    start = today - timedelta(days=N-1)
    date_map = { (start + timedelta(days=i)).isoformat(): 0.0 for i in range(N) }

    sales = Sale.query.filter(Sale.timestamp >= datetime.combine(start, datetime.min.time())).all()
    for s in sales:
        if s.timestamp:
            d = s.timestamp.date().isoformat()
            if d in date_map:
                date_map[d] += float(s.total or 0.0)

    labels = list(date_map.keys())
    data = [date_map[d] for d in labels]

    recent_sales = Sale.query.order_by(Sale.timestamp.desc()).limit(20).all()
    recent_customers = Customer.query.order_by(Customer.id.desc()).limit(20).all()

    recent_sales_html = '\n'.join([
        f"<li class='list-group-item'>#{s.id} — ₹{int(s.total or 0)} — {s.timestamp.strftime('%Y-%m-%d %H:%M')}</li>"
        for s in recent_sales
    ]) if recent_sales else '<li class="list-group-item">No sales</li>'

    recent_customers_html = '\n'.join([
        f"<li class='list-group-item'>{c.id} — {c.name} — {c.phone or ''}</li>"
        for c in recent_customers
    ]) if recent_customers else '<li class="list-group-item">No customers</li>'

    labels_json = json.dumps(labels)
    data_json = json.dumps(data)

    html = f'''
<!doctype html>
<html>
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width,initial-scale=1">
    <title>Carwash Dashboard</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet"/>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
</head>
<body class="p-3">
    <div class="container">
        <div class="d-flex justify-content-between align-items-center mb-3">
            <h3>Dashboard</h3>
            <div><a class="btn btn-sm btn-outline-secondary" href="/pos">Open POS</a> <a class="btn btn-sm btn-outline-primary" href="/export/all.xlsx">Export Excel</a></div>
        </div>

        <div class="row mb-3">
            <div class="col-md-3"><div class="card p-3">Customers<br><strong>{cust_count}</strong></div></div>
            <div class="col-md-3"><div class="card p-3">Vehicles<br><strong>{vehicle_count}</strong></div></div>
            <div class="col-md-3"><div class="card p-3">Services<br><strong>{service_count}</strong></div></div>
            <div class="col-md-3"><div class="card p-3">Sales<br><strong>{sale_count}</strong></div></div>
        </div>

        <div class="card mb-3 p-3">
            <h5>Sales (last {N} days)</h5>
            <canvas id="salesChart" height="80"></canvas>
        </div>

        <div class="row">
            <div class="col-md-6">
                <div class="card p-3 mb-3">
                    <h6>Recent Sales</h6>
                    <ul class="list-group">
                    {recent_sales_html}
                    </ul>
                </div>
            </div>
            <div class="col-md-6">
                <div class="card p-3 mb-3">
                    <h6>Recent Customers</h6>
                    <ul class="list-group">
                    {recent_customers_html}
                    </ul>
                </div>
            </div>
        </div>
    </div>

    <script>
        const labels = {labels_json};
        const data = {data_json};
        const ctx = document.getElementById('salesChart');
        new Chart(ctx, {{
            type: 'line',
            data: {{ labels: labels, datasets: [{{ label: 'Daily sales', data: data, borderColor: 'rgb(75, 192, 192)', tension: 0.2, fill: true }}] }},
            options: {{ scales: {{ x: {{ display: true }}, y: {{ beginAtZero: true }} }} }}
        }});
    </script>
</body>
</html>
'''

    return html


# --- Excel export (XLSX) ---
@app.route('/export/all.xlsx')
def export_all_xlsx():
    wb = Workbook()
    # remove default sheet
    default = wb.active
    wb.remove(default)

    # Customers
    ws = wb.create_sheet('Customers')
    ws.append(['id','name','phone','email'])
    for c in Customer.query.order_by(Customer.id).all():
        ws.append([c.id, c.name, c.phone or '', c.email or ''])

    # Vehicles
    ws = wb.create_sheet('Vehicles')
    ws.append(['id','reg_no','model','customer_id'])
    for v in Vehicle.query.order_by(Vehicle.id).all():
        ws.append([v.id, v.reg_no, v.model or '', v.customer_id or ''])

    # Services
    ws = wb.create_sheet('Services')
    ws.append(['id','name','price'])
    for s in Service.query.order_by(Service.id).all():
        ws.append([s.id, s.name, s.price])

    # Sales
    ws = wb.create_sheet('Sales')
    ws.append(['sale_id','customer_id','vehicle_id','total','method','timestamp'])
    for sale in Sale.query.order_by(Sale.id).all():
        ws.append([sale.id, sale.customer_id or '', sale.vehicle_id or '', sale.total or 0.0, sale.method or '', sale.timestamp.isoformat() if sale.timestamp else ''])

    # SaleItems
    ws = wb.create_sheet('SaleItems')
    ws.append(['id','sale_id','service_id','qty','price','line_total'])
    for si in SaleItem.query.order_by(SaleItem.id).all():
        ws.append([si.id, si.sale_id, si.service_id, si.qty, si.price, si.line_total])

    # Appointments
    ws = wb.create_sheet('Appointments')
    ws.append(['id','vehicle_id','service_id','scheduled_at','status','paid'])
    for a in Appointment.query.order_by(Appointment.id).all():
        ws.append([a.id, a.vehicle_id, a.service_id, a.scheduled_at.isoformat() if a.scheduled_at else '', a.status, bool(a.paid)])

    # Payments
    ws = wb.create_sheet('Payments')
    ws.append(['id','appointment_id','amount','method','timestamp'])
    for p in Payment.query.order_by(Payment.id).all():
        ws.append([p.id, p.appointment_id or '', p.amount, p.method or '', p.timestamp.isoformat() if p.timestamp else ''])

    # Save to bytes
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='carwash_export.xlsx')

# Run
if __name__ == '__main__':
    # initialize DB
    with app.app_context():
        init_db()
    # run server
    # For local development: use host='127.0.0.1'
    # For hosting/network access: use host='0.0.0.0'
    import os
    host = os.getenv('FLASK_HOST', '0.0.0.0')  # Default to 0.0.0.0 for hosting
    port = int(os.getenv('FLASK_PORT', 5000))
    debug = os.getenv('FLASK_DEBUG', 'False').lower() == 'true'
    
    try:
        print(f'Starting server on http://{host}:{port}')
        print(f'Access the POS at http://{host}:{port}/ or http://{host}:{port}/pos')
        app.run(host=host, port=port, debug=debug, use_reloader=debug)
    except Exception:
        import traceback, sys
        traceback.print_exc()
        print('\nServer failed to start. See traceback above for details.', file=sys.stderr)
        raise
